from openpyxl import load_workbook
from rest_framework import serializers

ALLOWED_STATUS = ['pending', 'in_progress', 'completed']

def read_excel_file(file):
    try:
        workbook = load_workbook(file, read_only=True)
        worksheet = workbook.active

        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows) 

        tasks = []

        for index, row in enumerate(rows, start=2):
            if not any(row):  # skip empty rows
                continue

            data = dict(zip(headers, row))
            if data['status'] not in ALLOWED_STATUS:
                continue
            tasks.append(data)

        return tasks
    except Exception as e:
        return None